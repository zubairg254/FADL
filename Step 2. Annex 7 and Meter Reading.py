import os
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk # Added ttk for progress bar
import pandas as pd
from datetime import datetime, timedelta
import re
from pathlib import Path
import PyPDF2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class FADLToExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Convert and Match FADL with Meter Reading")
        self.root.geometry("850x750") # Increased height for status/progress
        self.root.resizable(False, False)
        self.root.configure(bg='#f0f0f0')

        # File/folder path variables
        self.pdf_path = tk.StringVar()
        self.lp_folder_path = tk.StringVar()
        self.uch_excel_path = tk.StringVar()
        self.save_path = tk.StringVar()

        self.setup_gui()

    def setup_gui(self):
        main_frame = tk.Frame(self.root, bg='#f0f0f0', padx=20, pady=20)
        main_frame.pack(fill='both', expand=True)

        # Title
        title_label = tk.Label(main_frame, text="FADL to Excel Enhanced Tool", 
                              font=('Arial', 16, 'bold'), bg='#f0f0f0', fg='#333')
        title_label.pack(pady=(0, 20))

        # PDF Selection Frame
        pdf_frame = tk.Frame(main_frame, bg='#f0f0f0')
        pdf_frame.pack(fill='x', pady=(0, 10))
        tk.Label(pdf_frame, text="Select PDF File (for Annex 7):", font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(anchor='w')
        pdf_path_frame = tk.Frame(pdf_frame, bg='#f0f0f0')
        pdf_path_frame.pack(fill='x', pady=(5,0))
        tk.Entry(pdf_path_frame, textvariable=self.pdf_path, font=('Arial', 10), width=50).pack(side='left', fill='x', expand=True, padx=(0,10))
        tk.Button(pdf_path_frame, text="Browse", command=self.browse_pdf, bg='#4CAF50', fg='white', font=('Arial', 10), padx=10, pady=2).pack(side='right')

        # LP folder selection
        lp_frame = tk.Frame(main_frame, bg='#f0f0f0')
        lp_frame.pack(fill='x', pady=(0, 10))
        tk.Label(lp_frame, text="Select Folder with *.lp Files (for Meter Readings):", font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(anchor='w')
        lp_path_frame = tk.Frame(lp_frame, bg='#f0f0f0')
        lp_path_frame.pack(fill='x', pady=(5,0))
        tk.Entry(lp_path_frame, textvariable=self.lp_folder_path, font=('Arial', 10), width=50).pack(side='left', fill='x', expand=True, padx=(0,10))
        tk.Button(lp_path_frame, text="Browse", command=self.browse_lp_folder, bg='#4CAF50', fg='white', font=('Arial', 10), padx=10, pady=2).pack(side='right')

        # UCH Excel file selection
        uch_frame = tk.Frame(main_frame, bg='#f0f0f0')
        uch_frame.pack(fill='x', pady=(0, 10))
        tk.Label(uch_frame, text="Select Excel File with FADL_Calculation (for UCH Demand):", font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(anchor='w')
        uch_path_frame = tk.Frame(uch_frame, bg='#f0f0f0')
        uch_path_frame.pack(fill='x', pady=(5,0))
        tk.Entry(uch_path_frame, textvariable=self.uch_excel_path, font=('Arial', 10), width=50).pack(side='left', fill='x', expand=True, padx=(0,10))
        tk.Button(uch_path_frame, text="Browse", command=self.browse_uch_excel, bg='#4CAF50', fg='white', font=('Arial', 10), padx=10, pady=2).pack(side='right')

        # Excel save location
        excel_frame = tk.Frame(main_frame, bg='#f0f0f0')
        excel_frame.pack(fill='x', pady=(0, 15))
        tk.Label(excel_frame, text="Save Excel File As:", font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(anchor='w')
        excel_path_frame = tk.Frame(excel_frame, bg='#f0f0f0')
        excel_path_frame.pack(fill='x', pady=(5,0))
        tk.Entry(excel_path_frame, textvariable=self.save_path, font=('Arial', 10), width=50).pack(side='left', fill='x', expand=True, padx=(0,10))
        tk.Button(excel_path_frame, text="Browse", command=self.browse_save, bg='#2196F3', fg='white', font=('Arial', 10), padx=10, pady=2).pack(side='right')
        
        # Convert button
        tk.Button(main_frame, text="Convert to Excel", command=self.convert, width=20, bg='#FF9800', fg='white', font=('Arial', 12, 'bold'), padx=20, pady=10).pack(pady=20)

        # Progress Bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.pack(fill='x', pady=(0, 5))
        
        # Status label / Text Area
        self.status_text = tk.Text(main_frame, height=5, width=70, font=('Consolas', 9), bg='#f8f8f8', relief=tk.SOLID, borderwidth=1)
        self.status_text.pack(fill='both', expand=True, pady=(0,0))
        scrollbar = tk.Scrollbar(main_frame, command=self.status_text.yview)
        # scrollbar.pack(side='right', fill='y') # This can cause layout issues, let's see if text area scroll is enough
        self.status_text.config(yscrollcommand=scrollbar.set)
        self.update_status("Ready.")


    def browse_pdf(self):
        file = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF Files", "*.pdf"), ("All files", "*.*")]
        )
        if file:
            self.pdf_path.set(file)
            # Auto-suggest excel filename if save_path is empty
            if not self.save_path.get():
                pdf_p = Path(file)
                excel_name = pdf_p.parent / f"{pdf_p.stem}_FADL_converted.xlsx"
                self.save_path.set(str(excel_name))
            self.update_status(f"PDF selected: {file}")

    def browse_lp_folder(self):
        folder = filedialog.askdirectory(title="Select Folder with *.lp Files")
        if folder:
            self.lp_folder_path.set(folder)
            self.update_status(f"LP folder selected: {folder}")

    def browse_uch_excel(self):
        file = filedialog.askopenfilename(
            title="Select Excel File with FADL_Calculation",
            filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xls"), ("All files", "*.*")]
        )
        if file:
            self.uch_excel_path.set(file)
            self.update_status(f"UCH Excel file selected: {file}")

    def browse_save(self):
        # Suggest a filename based on PDF if available
        initial_file = ""
        if self.pdf_path.get():
            pdf_p = Path(self.pdf_path.get())
            initial_file = f"{pdf_p.stem}_FADL_converted.xlsx"
        
        file = filedialog.asksaveasfilename(
            title="Save Excel File As",
            initialfile=initial_file,
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx"), ("All files", "*.*")]
        )
        if file:
            self.save_path.set(file)
            self.update_status(f"Excel save location set: {file}")

    def update_status(self, message):
        self.status_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()

    def load_uch_excel_data(self, excel_path):
        """Load UCH Excel data from FADL_Calculation worksheet"""
        try:
            self.update_status(f"Loading UCH Excel data from: {excel_path}")
            # Load the FADL_Calculation worksheet
            df = pd.read_excel(excel_path, sheet_name='FADL_Calculation')
            
            # Check if required columns exist
            if 'Date' not in df.columns or 'Time to' not in df.columns:
                self.update_status("Error: Required columns 'Date' and 'Time to' not found in FADL_Calculation worksheet")
                return None
            
            # Get column E (index 4) - assuming it contains the UCH demand values
            if len(df.columns) < 5:
                self.update_status("Error: Column E not found in FADL_Calculation worksheet")
                return None
            
            # Create a lookup dictionary based on Date and Time to
            uch_lookup = {}
            for index, row in df.iterrows():
                date_val = row['Date']
                time_to_val = row['Time to']
                column_e_val = row.iloc[4]  # Column E (0-indexed, so index 4)
                
                # Handle different date formats
                if pd.isna(date_val) or pd.isna(time_to_val):
                    continue
                
                # Convert date to string format if it's a datetime object
                if isinstance(date_val, pd.Timestamp):
                    date_str = date_val.strftime('%d-%m-%Y')
                else:
                    date_str = str(date_val)
                
                # Convert time to string format if it's a time object
                if isinstance(time_to_val, pd.Timestamp):
                    time_str = time_to_val.strftime('%H:%M')
                else:
                    time_str = str(time_to_val)
                
                # Create lookup key
                lookup_key = f"{date_str}_{time_str}"
                uch_lookup[lookup_key] = column_e_val
            
            self.update_status(f"Successfully loaded {len(uch_lookup)} UCH demand records")
            return uch_lookup
            
        except Exception as e:
            self.update_status(f"Error loading UCH Excel data: {str(e)}")
            return None

    # --- PDF Processing Methods (to be integrated from PDFToExcelConverter) ---
    def extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF file"""
        self.update_status(f"Extracting text from PDF: {pdf_path}")
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for i, page in enumerate(pdf_reader.pages):
                    self.update_status(f"Reading page {i+1}/{len(pdf_reader.pages)}...")
                    text += page.extract_text() + "\n"
                self.update_status("Text extraction complete.")
                return text
        except Exception as e:
            self.update_status(f"Error reading PDF: {str(e)}")
            raise Exception(f"Error reading PDF: {str(e)}")

    def parse_pdf_data(self, text, meter_data_for_lookup=None, uch_lookup=None):
        """
        Parse the extracted text to extract structured data for Annex 7.
        Includes lookup for meter reading sum from meter_data_for_lookup, multiplied by 2.
        Also includes lookup for UCH demand from uch_lookup based on Date and Time To.
        """
        self.update_status("Parsing PDF data for Annex 7 and looking up Meter Readings...")
        
        lines = text.split('\n')
        data = []
        current_date = None
        
        date_pattern = r'(\d{2}-\d{2}-\d{4})'
        data_pattern = r'(\d{1,2}:\d{2})\s+(?:(\d{1,2}:\d{2})\s+)?(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+([0-9,]+\.\d+)\s*(\d+\.\d+)'
        simple_data_pattern = r'(\d{1,2}:\d{2})\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+([0-9,]+\.\d+)\s*(\d+\.\d+)'

        for line_idx, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            
            original_line_for_debug = line

            date_match = re.search(date_pattern, line)
            if date_match:
                current_date = date_match.group(1)
                line = line[date_match.end():].strip()

            match = re.search(data_pattern, line)
            if not match:
                match = re.search(simple_data_pattern, line)
                if match:
                    groups = match.groups()
                    match = lambda: None 
                    match.groups = lambda: (groups[0], None) + groups[1:]

            if match and current_date:
                try:
                    g = match.groups()
                    time_from_str = g[0]
                    time_to_str = g[1] if g[1] else ""

                    if not time_to_str and time_from_str:
                        try:
                            if current_date:
                                datetime_from_str = f"{current_date} {time_from_str}"
                                dt_from_obj = datetime.strptime(datetime_from_str, "%d-%m-%Y %H:%M")
                                dt_to_obj = dt_from_obj + timedelta(minutes=30)
                                time_to_str = dt_to_obj.strftime("%H:%M")
                        except ValueError as ve:
                            self.update_status(f"Warning: Could not parse Time From '{time_from_str}' with date '{current_date}' to infer Time To (30 min rule): {ve}")

                    wapda_demand = float(g[2])
                    level_achieved = float(g[3])
                    tolerance = float(g[4])
                    non_compliance = float(g[5])
                    amount_str = g[6]
                    amount = float(amount_str.replace(',', ''))
                    rate = float(g[7])

                    excel_formula_string = pd.NA # Default to Not Available
                    # Ensure time_to_str is valid before attempting lookup
                    if meter_data_for_lookup is not None and not meter_data_for_lookup.empty and current_date and time_to_str: 
                        try:
                            lookup_timestamp_str = f"{current_date} {time_to_str}" # CHANGED to time_to_str
                            lookup_timestamp = datetime.strptime(lookup_timestamp_str, "%d-%m-%Y %H:%M")
                            
                            if lookup_timestamp in meter_data_for_lookup.index:
                                meter_values = []
                                # Get all individual meter columns (exclude 'Sum' if it's there, though it shouldn't be used for this formula)
                                individual_meter_columns = [col for col in meter_data_for_lookup.columns if col.lower() != 'sum']
                                
                                for col_name in individual_meter_columns:
                                    val = meter_data_for_lookup.loc[lookup_timestamp, col_name]
                                    # Treat NA or non-numeric as 0 for formula construction
                                    if pd.isna(val) or not isinstance(val, (int, float)):
                                        meter_values.append("0")
                                    else:
                                        meter_values.append(str(val))
                                
                                if meter_values:
                                    formula_core = "+".join(meter_values)
                                    excel_formula_string = f"=({formula_core})*2"
                                # else: excel_formula_string remains pd.NA if no meter columns or values found
                                    
                        except ValueError as ve_lookup:
                            self.update_status(f"Timestamp format error for lookup: {lookup_timestamp_str} -> {ve_lookup}")
                        except Exception as ex_lookup:
                            self.update_status(f"Error during meter reading lookup for formula construction {lookup_timestamp_str}: {ex_lookup}")
                    
                    # UCH Demand lookup
                    uch_demand_value = pd.NA
                    if uch_lookup is not None and current_date and time_to_str:
                        lookup_key = f"{current_date}_{time_to_str}"
                        if lookup_key in uch_lookup:
                            uch_demand_value = uch_lookup[lookup_key]
                    
                    data.append({
                        'Date': current_date, 'Time From': time_from_str, 'Time To': time_to_str,
                        'WAPDA Demand MW': wapda_demand, 'Level Achieved MW': level_achieved,
                        'Meter Reading Sum': excel_formula_string, # Store formula string
                        'Demand Calculated by UCH': uch_demand_value,
                        'Tolerance MW': tolerance, 'Non-Compliance MWh': non_compliance,
                        'Rate Rs./kWh': rate, 'Amount Rs.': amount
                    })
                except Exception as e:
                    self.update_status(f"Skipping line (parse error): {original_line_for_debug} -> {e}")
                    continue
            elif current_date and not re.match(r'^[A-Za-z]', line) and len(line.split()) > 5:
                parts = line.split()
                try:
                    if ':' in parts[0]:
                        time_from_str_fb = parts[0]
                        time_to_str_fb = parts[1] if len(parts) > 1 and ':' in parts[1] else ""
                        
                        if not time_to_str_fb and time_from_str_fb and current_date:
                            try:
                                datetime_from_str_fb = f"{current_date} {time_from_str_fb}"
                                dt_from_obj_fb = datetime.strptime(datetime_from_str_fb, "%d-%m-%Y %H:%M")
                                dt_to_obj_fb = dt_from_obj_fb + timedelta(minutes=30)
                                time_to_str_fb = dt_to_obj_fb.strftime("%H:%M")
                            except ValueError: pass

                        num_parts_offset = 2 if time_to_str_fb and ':' in parts[1] else 1
                        
                        excel_formula_string_fb = pd.NA
                        # Ensure time_to_str_fb is valid before attempting lookup
                        if meter_data_for_lookup is not None and not meter_data_for_lookup.empty and current_date and time_to_str_fb: 
                            try:
                                lookup_timestamp_str_fb = f"{current_date} {time_to_str_fb}" # CHANGED to time_to_str_fb
                                lookup_timestamp_fb = datetime.strptime(lookup_timestamp_str_fb, "%d-%m-%Y %H:%M")
                                if lookup_timestamp_fb in meter_data_for_lookup.index:
                                    meter_values_fb = []
                                    individual_meter_columns_fb = [col for col in meter_data_for_lookup.columns if col.lower() != 'sum']
                                    for col_name_fb in individual_meter_columns_fb:
                                        val_fb = meter_data_for_lookup.loc[lookup_timestamp_fb, col_name_fb]
                                        if pd.isna(val_fb) or not isinstance(val_fb, (int, float)):
                                            meter_values_fb.append("0")
                                        else:
                                            meter_values_fb.append(str(val_fb))
                                    if meter_values_fb:
                                        formula_core_fb = "+".join(meter_values_fb)
                                        excel_formula_string_fb = f"=({formula_core_fb})*2"
                            except (ValueError, KeyError, AttributeError): pass
                        
                        # UCH Demand lookup for fallback
                        uch_demand_value_fb = pd.NA
                        if uch_lookup is not None and current_date and time_to_str_fb:
                            lookup_key_fb = f"{current_date}_{time_to_str_fb}"
                            if lookup_key_fb in uch_lookup:
                                uch_demand_value_fb = uch_lookup[lookup_key_fb]
                        
                        if len(parts) >= num_parts_offset + 6:
                            wapda_demand = float(parts[num_parts_offset])
                            level_achieved = float(parts[num_parts_offset+1])
                            tolerance = float(parts[num_parts_offset+2])
                            non_compliance = float(parts[num_parts_offset+3])
                            rate_val = float(parts[num_parts_offset+4])
                            amount_val = float(parts[num_parts_offset+5].replace(',', ''))
                            data.append({
                                'Date': current_date, 'Time From': time_from_str_fb, 'Time To': time_to_str_fb,
                                'WAPDA Demand MW': wapda_demand, 'Level Achieved MW': level_achieved,
                                'Meter Reading Sum': excel_formula_string_fb,
                                'Demand Calculated by UCH': uch_demand_value_fb,
                                'Tolerance MW': tolerance, 'Non-Compliance MWh': non_compliance,
                                'Rate Rs./kWh': rate_val, 'Amount Rs.': amount_val
                            })
                        elif len(parts) >= num_parts_offset + 5:
                             wapda_demand = float(parts[num_parts_offset])
                             level_achieved = float(parts[num_parts_offset+1])
                             tolerance = float(parts[num_parts_offset+2])
                             non_compliance = float(parts[num_parts_offset+3])
                             amount_val = float(parts[num_parts_offset+4].replace(',', ''))
                             data.append({
                                'Date': current_date, 'Time From': time_from_str_fb, 'Time To': time_to_str_fb,
                                'WAPDA Demand MW': wapda_demand, 'Level Achieved MW': level_achieved,
                                'Meter Reading Sum': excel_formula_string_fb, # Still uses the same looked up value/formula
                                'Demand Calculated by UCH': uch_demand_value_fb,
                                'Tolerance MW': tolerance, 'Non-Compliance MWh': non_compliance,
                                'Rate Rs./kWh': 6.0534, 'Amount Rs.': amount_val
                            })
                except (ValueError, IndexError): pass

        if not data:
            self.update_status("No structured data found in PDF for Annex 7. Trying a simpler line-by-line scan.")
            for line in lines:
                if re.search(r'\d{1,2}:\d{2}', line) and len(line.split()) >= 6:
                    parts = line.split()
                    numbers = []
                    time_parts = []
                    for part in parts:
                        if ':' in part and len(part) <= 5: time_parts.append(part)
                        else:
                            try: numbers.append(float(part.replace(',', '')))
                            except ValueError: continue
                    
                    if len(numbers) >= 5 and len(time_parts) >= 1:
                        time_from_fb2 = time_parts[0] if time_parts else ""
                        time_to_fb2 = time_parts[1] if len(time_parts) > 1 else ""

                        if not time_to_fb2 and time_from_fb2 and current_date and current_date != "Unknown":
                            try:
                                datetime_from_str_fb2 = f"{current_date} {time_from_fb2}"
                                dt_from_obj_fb2 = datetime.strptime(datetime_from_str_fb2, "%d-%m-%Y %H:%M")
                                dt_to_obj_fb2 = dt_from_obj_fb2 + timedelta(minutes=30)
                                time_to_fb2 = dt_to_obj_fb2.strftime("%H:%M")
                            except ValueError: pass
                        
                        excel_formula_string_fb2 = pd.NA
                        # Ensure time_to_fb2 is valid for lookup
                        if meter_data_for_lookup is not None and not meter_data_for_lookup.empty and \
                           current_date and current_date != "Unknown" and time_to_fb2: 
                            try:
                                lookup_timestamp_str_fb2 = f"{current_date} {time_to_fb2}" # CHANGED to time_to_fb2
                                lookup_timestamp_fb2 = datetime.strptime(lookup_timestamp_str_fb2, "%d-%m-%Y %H:%M")
                                if lookup_timestamp_fb2 in meter_data_for_lookup.index:
                                    meter_values_fb2 = []
                                    individual_meter_columns_fb2 = [col for col in meter_data_for_lookup.columns if col.lower() != 'sum']
                                    for col_name_fb2 in individual_meter_columns_fb2:
                                        val_fb2 = meter_data_for_lookup.loc[lookup_timestamp_fb2, col_name_fb2]
                                        if pd.isna(val_fb2) or not isinstance(val_fb2, (int, float)):
                                            meter_values_fb2.append("0")
                                        else:
                                            meter_values_fb2.append(str(val_fb2))
                                    if meter_values_fb2:
                                        formula_core_fb2 = "+".join(meter_values_fb2)
                                        excel_formula_string_fb2 = f"=({formula_core_fb2})*2"
                            except (ValueError, KeyError, AttributeError): pass

                        # UCH Demand lookup for final fallback
                        uch_demand_value_fb2 = pd.NA
                        if uch_lookup is not None and current_date and current_date != "Unknown" and time_to_fb2:
                            lookup_key_fb2 = f"{current_date}_{time_to_fb2}"
                            if lookup_key_fb2 in uch_lookup:
                                uch_demand_value_fb2 = uch_lookup[lookup_key_fb2]

                        data.append({
                            'Date': current_date or "Unknown", 'Time From': time_from_fb2, 'Time To': time_to_fb2,
                            'WAPDA Demand MW': numbers[0] if len(numbers) > 0 else 0,
                            'Level Achieved MW': numbers[1] if len(numbers) > 1 else 0,
                            'Meter Reading Sum': excel_formula_string_fb2,
                            'Demand Calculated by UCH': uch_demand_value_fb2,
                            'Tolerance MW': numbers[2] if len(numbers) > 2 else 0,
                            'Non-Compliance MWh': numbers[3] if len(numbers) > 3 else 0,
                            'Rate Rs./kWh': 6.0534, 
                            'Amount Rs.': numbers[4] if len(numbers) > 4 else 0
                        })
                        if not current_date: self.update_status("Warning: PDF data parsed without a clear date context for some rows.")

        if data:
             self.update_status(f"Successfully parsed {len(data)} rows for Annex 7.")
        else:
            self.update_status("Could not parse any data for Annex 7 from the PDF.")
        return pd.DataFrame(data)


    def write_annex7_sheet(self, writer, df_annex7):
        """Writes the Annex 7 data to a sheet with formatting."""
        self.update_status("Writing 'Annex 7' sheet...")
        if df_annex7.empty:
            self.update_status("'Annex 7' data is empty. Skipping sheet writing.")
            # Create an empty sheet or a sheet with a message
            empty_df = pd.DataFrame([{"Message": "No data extracted from PDF for Annex 7."}])
            empty_df.to_excel(writer, sheet_name="Annex 7", index=False)
            return

        # Ensure DataFrame columns are in the desired order before writing
        # This helps if parse_pdf_data might sometimes miss a column, though it should create all.
        expected_cols_order = [
            "Date", "Time From", "Time To", "WAPDA Demand MW", "Level Achieved MW", 
            "Meter Reading Sum", "Demand Calculated by UCH", "Tolerance MW", "Non-Compliance MWh", "Rate Rs./kWh", "Amount Rs."
        ]
        # If df_annex7 is missing some of these, reindex will add them with NAs.
        df_annex7_ordered = df_annex7.reindex(columns=expected_cols_order)

        df_annex7_ordered.to_excel(writer, sheet_name="Annex 7", index=False)
        worksheet = writer.sheets["Annex 7"]

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet[1]: # First row for headers
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Column widths and number formats
        # New order: Date, Time From, Time To, WAPDA Demand MW, Level Achieved MW, Meter Reading Sum, Demand Calculated by UCH, Tolerance MW, Non-Compliance MWh, Rate Rs./kWh, Amount Rs.
        column_settings = {
            'A': {'width': 12},                                      # Date
            'B': {'width': 10},                                      # Time From
            'C': {'width': 10},                                      # Time To
            'D': {'width': 18, 'format': '#,##0.00'},                # WAPDA Demand MW
            'E': {'width': 18, 'format': '#,##0.00'},                # Level Achieved MW
            'F': {'width': 18, 'format': '#,##0.00'},                # Meter Reading Sum
            'G': {'width': 22, 'format': '#,##0.00'},                # Demand Calculated by UCH (NEW)
            'H': {'width': 15, 'format': '#,##0.00'},                # Tolerance MW
            'I': {'width': 20, 'format': '#,##0.00'},                # Non-Compliance MWh
            'J': {'width': 15, 'format': '#,##0.0000'},              # Rate Rs./kWh
            'K': {'width': 18, 'format': '#,##0.00'}                 # Amount Rs.
        }

        # Ensure DataFrame columns match the expected order for column_settings
        # This is important if pd.NA values caused a column to be all NA and potentially affect its type or presence for openpyxl
        expected_cols_order = [
            "Date", "Time From", "Time To", "WAPDA Demand MW", "Level Achieved MW", 
            "Meter Reading Sum", "Demand Calculated by UCH", "Tolerance MW", "Non-Compliance MWh", "Rate Rs./kWh", "Amount Rs."
        ]
        # Reindex df_annex7 to ensure columns exist and are in order, fill missing with pd.NA
        # This is more of a safeguard; parse_pdf_data should ideally create all columns.
        df_annex7_reordered = df_annex7.reindex(columns=expected_cols_order)
        
        # Write the potentially reordered DataFrame
        df_annex7_reordered.to_excel(writer, sheet_name="Annex 7", index=False)
        worksheet = writer.sheets["Annex 7"] # Re-assign worksheet after writing potentially new df

        # Re-apply header formats as writing df_annex7_reordered might clear them
        for cell in worksheet[1]: 
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for col_idx, col_letter in enumerate(column_settings.keys()):
            settings = column_settings[col_letter]
            worksheet.column_dimensions[col_letter].width = settings['width']
            if 'format' in settings:
                # Apply format to all cells in the column except the header
                for row_idx in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                    # Check if cell has a value, as formatting an empty cell is useless 
                    # and pd.NA might be written as empty or a specific string by to_excel
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value == str(pd.NA)):
                        cell.number_format = settings['format']
        
        self.update_status("'Annex 7' sheet written and formatted.")


    # --- LP File Processing Methods (existing) ---
    def process_lp_files_folder(self, folder_path):
        """
        Process all *.lp files in the folder, extract 2.9 MW readings and timestamps,
        align by timestamp, and return a DataFrame with sum.
        """
        self.update_status(f"Processing LP files from: {folder_path}")
        lp_files = glob.glob(os.path.join(folder_path, "*.lp"))
        if not lp_files:
            self.update_status("No *.lp files found in the selected folder.")
            # Return an empty DataFrame or raise error based on desired behavior
            return pd.DataFrame() # Return empty DF to avoid breaking Excel writing

        meter_dfs = []
        meter_names = []
        for lp_file in lp_files:
            meter_name = os.path.splitext(os.path.basename(lp_file))[0]
            self.update_status(f"Parsing LP file: {meter_name}.lp")
            try:
                df = self.parse_lp_file(lp_file)
                if df is not None and not df.empty:
                    df = df[['Timestamp', '2.9']].rename(columns={'2.9': meter_name})
                    meter_dfs.append(df)
                    meter_names.append(meter_name)
                else:
                    self.update_status(f"No data extracted from {meter_name}.lp or file was empty.")
            except Exception as e:
                self.update_status(f"Error parsing LP file {meter_name}.lp: {e}")
                # Optionally, continue with other files or re-raise

        if not meter_dfs:
            self.update_status("No valid meter data found in any *.lp files.")
            return pd.DataFrame()

        self.update_status("Merging data from all LP files...")
        merged_df = meter_dfs[0]
        for df_idx in range(1, len(meter_dfs)):
            merged_df = pd.merge(merged_df, meter_dfs[df_idx], on='Timestamp', how='outer')

        # Convert Timestamp to datetime objects if not already
        merged_df['Timestamp'] = pd.to_datetime(merged_df['Timestamp'])
        merged_df.sort_values('Timestamp', inplace=True)
        merged_df.reset_index(drop=True, inplace=True)

        meter_cols = [name for name in meter_names if name in merged_df.columns] # Ensure columns exist
        if meter_cols: # only sum if there are meter columns
            merged_df['Sum'] = merged_df[meter_cols].sum(axis=1, skipna=True)
        else:
            merged_df['Sum'] = 0 # Or handle as error/warning
        
        # Keep Timestamp as datetime objects for Excel, it handles formatting better.
        # The string conversion will be done by Excel's number formatting.
        # merged_df['Timestamp'] = merged_df['Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S') 
        self.update_status("LP file processing complete.")
        return merged_df

    def parse_lp_file(self, filepath):
        """
        Adapted Uch 1 logic: Extracts Timestamp and 2.9 MW readings.
        """
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        data = []
        current_ts = None
        for line in lines:
            line = line.strip()
            if line.startswith("P.01("):
                try:
                    ts_str = line.split('(')[1].split(')')[0]
                    # Ensure ts_str is exactly 12 characters for "%y%m%d%H%M%S"
                    if len(ts_str) == 12:
                         current_ts = datetime.strptime(ts_str, "%y%m%d%H%M%S")
                    elif len(ts_str) == 14: # Some formats might include seconds, though pattern is for 12
                         current_ts = datetime.strptime(ts_str, "%Y%m%d%H%M%S") # YY vs YYYY
                    else:
                        # self.update_status(f"Warning: Unusual timestamp string length in {os.path.basename(filepath)}: {ts_str}")
                        current_ts = None # Skip if format is unexpected
                        continue
                except ValueError as e:
                    # self.update_status(f"Warning: Timestamp parsing error in {os.path.basename(filepath)} for '{ts_str}': {e}")
                    current_ts = None
            elif current_ts and line and line.startswith("(") and line.endswith(")"):
                try:
                    # Expecting format like (00000001)(00000002.9)(00000003)...
                    # We need the value associated with "2.9 MW", which is stated as the second value in the block.
                    # The problem description implies the second value is "2.9", not that the value itself is 2.9.
                    # Let's assume "2.9" is a label for the second value in the reading block.
                    
                    # Split by ')(', then clean up parentheses
                    parts = line[1:-1].split(')(')
                    
                    if len(parts) >= 2: # Need at least two values for the "2.9 MW" reading (index 1)
                        # The values are floats, according to the original code.
                        # The problem implies the '2.9' is an identifier for the *column* or *type* of reading
                        # and that its value is the second numeric part.
                        value_for_2_9_mw = float(parts[1])
                        data.append([current_ts, value_for_2_9_mw])
                        current_ts += timedelta(minutes=30) # Increment for the next expected reading
                    else:
                        # self.update_status(f"Warning: Not enough data parts in line: {line} in {os.path.basename(filepath)}")
                        pass

                except (ValueError, IndexError) as e:
                    # self.update_status(f"Warning: Data parsing error in line: {line} in {os.path.basename(filepath)}: {e}")
                    continue # Skip malformed data line
        
        if data:
            df = pd.DataFrame(data, columns=['Timestamp', '2.9'])
            return df
        else:
            # self.update_status(f"No data parsed from LP file: {os.path.basename(filepath)}")
            return pd.DataFrame(columns=['Timestamp', '2.9']) # Ensure consistent return type

    # --- Convert Method (to be updated) ---
    def convert(self):
        # This method will be filled in the next step
        self.update_status("Starting conversion process...")
        self.progress.start()

        pdf_file = self.pdf_path.get()
        lp_folder = self.lp_folder_path.get()
        uch_excel_file = self.uch_excel_path.get()
        excel_file = self.save_path.get()

        if not excel_file:
            messagebox.showerror("Error", "Please specify Excel save location.")
            self.progress.stop()
            self.update_status("Conversion failed: No save location.")
            return

        # Process PDF for Annex 7
        df_annex7 = pd.DataFrame() # Initialize as empty
        if pdf_file:
            try:
                pdf_text = self.extract_text_from_pdf(pdf_file)
                if pdf_text:
                    df_annex7 = self.parse_pdf_data(pdf_text)
                else:
                    self.update_status("PDF text extraction failed or returned empty.")
                    # df_annex7 remains empty, write_annex7_sheet will handle this
            except Exception as e:
                messagebox.showerror("PDF Error", f"Error processing PDF file: {e}")
                self.update_status(f"PDF processing error: {e}")
                # Decide if to continue or stop; for now, let's try to process LP if available
        else:
            self.update_status("No PDF file selected. 'Annex 7' sheet will be empty or contain a message.")


        # Process LP files for Meter Readings
        meter_df = pd.DataFrame() # Initialize as empty
        meter_df_indexed = None # For meter data lookup
        if lp_folder:
            try:
                meter_df = self.process_lp_files_folder(lp_folder) # Returns df with Timestamp as datetime
                if not meter_df.empty:
                    # Prepare for lookup: Set Timestamp as index
                    # Ensure Timestamp is definite datetime before setting as index
                    meter_df['Timestamp'] = pd.to_datetime(meter_df['Timestamp'])
                    meter_df_indexed = meter_df.set_index('Timestamp')
                    self.update_status("Meter Reading data indexed for lookup.")
                else:
                    self.update_status("No data processed from LP files. Lookup will not be available.")
            except Exception as e:
                messagebox.showerror("LP File Error", f"Error processing LP files: {e}")
                self.update_status(f"LP file processing error: {e}")
        else:
            self.update_status("No LP folder selected. Meter Reading lookup will not be available.")

        # Load UCH Excel data
        uch_lookup = None
        if uch_excel_file:
            try:
                uch_lookup = self.load_uch_excel_data(uch_excel_file)
                if uch_lookup is None:
                    self.update_status("UCH Excel data could not be loaded. UCH demand lookup will not be available.")
            except Exception as e:
                messagebox.showerror("UCH Excel Error", f"Error loading UCH Excel file: {e}")
                self.update_status(f"UCH Excel loading error: {e}")
        else:
            self.update_status("No UCH Excel file selected. UCH demand lookup will not be available.")

        # Process PDF for Annex 7 - now pass meter_df_indexed and uch_lookup for lookup
        df_annex7 = pd.DataFrame() 
        if pdf_file:
            try:
                pdf_text = self.extract_text_from_pdf(pdf_file)
                if pdf_text:
                    # Pass the indexed meter data and UCH lookup to parse_pdf_data
                    df_annex7 = self.parse_pdf_data(pdf_text, meter_df_indexed, uch_lookup)
                else:
                    self.update_status("PDF text extraction failed or returned empty.")
            except Exception as e:
                messagebox.showerror("PDF Error", f"Error processing PDF file: {e}")
                self.update_status(f"PDF processing error: {e}")
        else:
            self.update_status("No PDF file selected. 'Annex 7' sheet will be empty or contain a message.")


        # Write to Excel
        self.update_status(f"Writing data to Excel: {excel_file}")
        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # Write "Annex 7"
                self.write_annex7_sheet(writer, df_annex7)

                # Write Meter Reading sheet
                if not meter_df.empty:
                    meter_df.to_excel(writer, sheet_name="Meter Reading", index=False)
                    worksheet_mr = writer.sheets["Meter Reading"]
                    # Apply datetime formatting to the Timestamp column (Column A)
                    # Standard Excel format for date and time
                    datetime_format = 'yyyy-mm-dd hh:mm:ss' 
                    # Adjust column width for Timestamp
                    worksheet_mr.column_dimensions['A'].width = 20 
                    for row in worksheet_mr.iter_rows(min_row=2, min_col=1, max_col=1): # Iterate over column A, skip header
                        for cell in row:
                            if cell.value is not None: # Ensure cell is not empty
                                cell.number_format = datetime_format
                    self.update_status("'Meter Reading' sheet written and Timestamp column formatted.")
                else:
                    # Create an empty sheet or a sheet with a message for Meter Reading
                    empty_meter_df = pd.DataFrame([{"Message": "No data processed from LP files." if lp_folder else "LP folder not selected."}])
                    empty_meter_df.to_excel(writer, sheet_name="Meter Reading", index=False)
                    self.update_status("'Meter Reading' sheet: No data.")

            self.update_status("Excel file created successfully!")
            messagebox.showinfo("Success", f"Excel file created:\n{excel_file}")
        except Exception as e:
            self.update_status(f"Error writing Excel file: {e}")
            messagebox.showerror("Excel Error", f"Error writing Excel file: {e}")
        finally:
            self.progress.stop()
            self.update_status("Conversion process finished.")


if __name__ == "__main__":
    root = tk.Tk()
    app = FADLToExcelApp(root)
    root.mainloop()
