import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np # For pd.NA / np.nan
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import threading
import datetime
from dateutil.relativedelta import relativedelta
import os
import subprocess
import sys

class LoadProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FADL Load Processor - New Logic") 
        self.root.geometry("875x500") 
        self.root.minsize(600, 500)

        self.file_path_var = tk.StringVar()
        self.month_var = tk.StringVar()
        self.start_load_type_var = tk.StringVar(value="FinalAvailabilityHourly") 
        self.custom_load_var = tk.DoubleVar(value=0.0)
        self.export_dir_var = tk.StringVar()
        self.df_dispatch = None
        self.df_availability = None
        self.availability_series_hourly_lookup = None 

        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)

        file_frame = ttk.LabelFrame(main_frame, text="Input File", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=5, pady=5)
        file_frame.columnconfigure(1, weight=1)
        ttk.Label(file_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(file_frame, textvariable=self.file_path_var, width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(file_frame, text="Browse...", command=self.browse_file).grid(row=0, column=2, sticky=tk.E, padx=5, pady=5)

        config_frame = ttk.LabelFrame(main_frame, text="Processing Configuration", padding="10")
        config_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
        config_frame.columnconfigure(1, weight=1)
        ttk.Label(config_frame, text="Month to Process:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.month_combo = ttk.Combobox(config_frame, textvariable=self.month_var, state="readonly", width=15)
        self.month_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        self.month_combo['values'] = ["Select a file first"]
        ttk.Label(config_frame, text="Starting Load Type:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        final_avail_radio = ttk.Radiobutton(config_frame, 
                                            text="Use Final Availability (hourly from Col D)", 
                                            variable=self.start_load_type_var, 
                                            value="FinalAvailabilityHourly", 
                                            command=self.toggle_custom_load_entry)
        final_avail_radio.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        custom_load_frame = ttk.Frame(config_frame) 
        custom_load_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=0, pady=0) 
        custom_radio = ttk.Radiobutton(custom_load_frame, 
                                       text="Custom Load:", 
                                       variable=self.start_load_type_var, 
                                       value="Custom", 
                                       command=self.toggle_custom_load_entry)
        custom_radio.grid(row=0, column=0, sticky=tk.W, pady=2)
        self.custom_load_entry = ttk.Entry(custom_load_frame, 
                                           textvariable=self.custom_load_var, 
                                           width=10, 
                                           state=tk.DISABLED) 
        self.custom_load_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)
        
        export_frame = ttk.LabelFrame(main_frame, text="Output", padding="10")
        export_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5) 
        export_frame.columnconfigure(1, weight=1)
        ttk.Label(export_frame, text="Export Folder:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(export_frame, textvariable=self.export_dir_var, width=40).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(export_frame, text="Browse...", command=self.browse_export_dir).grid(row=0, column=2, sticky=tk.E, padx=5, pady=5)

        run_frame = ttk.Frame(main_frame, padding="10")
        run_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=5, pady=5)
        run_frame.columnconfigure(0, weight=1) 
        self.process_button = ttk.Button(run_frame, text="Start Processing", command=self.start_processing_thread, width=20)
        self.process_button.grid(row=0, column=1, sticky=tk.E, padx=5, pady=5)
        self.progress_bar = ttk.Progressbar(run_frame, orient=tk.HORIZONTAL, mode='determinate')
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)
        
        log_frame = ttk.LabelFrame(main_frame, text="Status Log", padding="10")
        log_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1) 
        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD, state=tk.DISABLED)
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text['yscrollcommand'] = log_scroll.set
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        self.status_log("Application started. Please select an Excel file.")
        self.toggle_custom_load_entry() 

    def browse_file(self):
        file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*")))
        if file_path:
            self.file_path_var.set(file_path)
            self.status_log(f"Selected input file: {file_path}")
            self.df_dispatch, self.df_availability, self.availability_series_hourly_lookup = None, None, None
            self.process_button.config(state=tk.DISABLED) 
            _, _, available_months = self.read_excel_data(file_path) 
            if self.df_dispatch is not None and available_months: 
                self.month_combo['values'] = available_months
                self.month_var.set(available_months[0] if available_months else "")
                self.status_log(f"Available months populated: {available_months if available_months else 'None'}")
                if self.df_availability is not None : 
                    self.process_button.config(state=tk.NORMAL) 
                    self._prepare_hourly_availability_lookup()
            else:
                self.month_combo['values'] = ["Error reading/parsing file"]
                self.month_var.set("")
                self.status_log("Could not populate months. Check file/sheet integrity and Dispatch Col A for dates.")
        else: self.status_log("File selection cancelled.")

    def browse_export_dir(self):
        dir_path = filedialog.askdirectory(title="Select Export Directory")
        if dir_path: self.export_dir_var.set(dir_path); self.status_log(f"Selected export directory: {dir_path}")
        else: self.status_log("Export directory selection cancelled.")

    def toggle_custom_load_entry(self):
        self.custom_load_entry.config(state=tk.NORMAL if self.start_load_type_var.get() == "Custom" else tk.DISABLED)
        if self.start_load_type_var.get() != "Custom": self.custom_load_var.set(0.0)
        self.status_log(f"Start load type set to: {self.start_load_type_var.get()}")

    def _prepare_hourly_availability_lookup(self):
        self.availability_series_hourly_lookup = None 
        if self.df_availability is None or self.df_availability.empty:
            self.root.after(0, self.status_log_safe,"Availability data empty, cannot prepare hourly lookup.") # Use root.after for thread safety
            return
        try:
            if len(self.df_availability.columns) < 4:
                self.root.after(0, self.status_log_safe,"Availability sheet < 4 columns. Cannot prepare hourly lookup from Col D.")
                return
            avail_df = self.df_availability.copy()
            if not pd.api.types.is_datetime64_any_dtype(avail_df.iloc[:, 0]):
                avail_df.iloc[:, 0] = pd.to_datetime(avail_df.iloc[:, 0], errors='coerce')
            if not pd.api.types.is_numeric_dtype(avail_df.iloc[:, 3]): 
                avail_df.iloc[:, 3] = pd.to_numeric(avail_df.iloc[:, 3], errors='coerce')
            avail_df.dropna(subset=[avail_df.columns[0], avail_df.columns[3]], inplace=True)
            if avail_df.empty:
                self.root.after(0, self.status_log_safe,"No valid data in Availability sheet (Cols A & D) for hourly lookup.")
                return
            avail_df = avail_df.set_index(avail_df.columns[0])
            load_series = avail_df.iloc[:, 2] 
            self.availability_series_hourly_lookup = load_series.resample('h').first()
            self.root.after(0, self.status_log_safe,"Hourly availability lookup table prepared.")
        except Exception as e:
            self.root.after(0, self.status_log_safe,f"Error preparing hourly availability lookup: {e}")
            self.availability_series_hourly_lookup = None

    def get_hourly_final_availability(self, timestamp_obj): # Removed unused availability_df_month
        if self.availability_series_hourly_lookup is None:
            # self.root.after(0, self.status_log_safe, f"Warning: Hourly availability lookup series not prepared. Cannot get value for {timestamp_obj}.") # Potentially too verbose
            return None
        try:
            lookup_hour = timestamp_obj.replace(minute=0, second=0, microsecond=0)
            available_load = self.availability_series_hourly_lookup.get(lookup_hour)
            return float(available_load) if pd.notna(available_load) else None
        except Exception as e:
            self.root.after(0, self.status_log_safe, f"Error during hourly availability lookup for {timestamp_obj}: {e}")
            return None

    def read_excel_data(self, file_path):
        # ... (read_excel_data content remains largely the same, ensure it uses self.status_log directly)
        self.status_log(f"Reading Excel file: {file_path}")
        self.df_dispatch = None
        self.df_availability = None
        available_months = []
        if not file_path:
            messagebox.showerror("Error", "No file selected to read.")
            self.status_log("File reading skipped: No file path provided.")
            return None, None, []
        try:
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names
            required_sheets = {"Dispatch Instructions", "Availability"}
            if not required_sheets.issubset(sheet_names):
                missing_sheets = required_sheets - set(sheet_names)
                err_msg = f"Missing required sheet(s): {', '.join(missing_sheets)}."
                messagebox.showerror("Sheet Error", err_msg + f"\nFound sheets: {sheet_names}")
                self.status_log(f"Error: {err_msg} Excel file only contains: {sheet_names}")
                return None, None, []
            self.status_log(f"Found sheets: {sheet_names}. Reading 'Dispatch Instructions' and 'Availability'.")
            self.df_dispatch = pd.read_excel(xls, sheet_name="Dispatch Instructions")
            self.status_log(f"'Dispatch Instructions' sheet read. Rows: {len(self.df_dispatch)}, Columns: {len(self.df_dispatch.columns) if self.df_dispatch is not None else 0}")
            if self.df_dispatch is None or self.df_dispatch.empty: 
                self.status_log("Warning: 'Dispatch Instructions' sheet is empty or failed to load. No months to process.")
                self.df_dispatch = None 
            elif self.df_dispatch.columns.empty:
                 self.status_log("Warning: 'Dispatch Instructions' sheet has no columns.")
                 messagebox.showwarning("Data Warning", "'Dispatch Instructions' sheet has no columns.")
                 self.df_dispatch = None 
            elif not pd.api.types.is_datetime64_any_dtype(self.df_dispatch.iloc[:, 0]): 
                self.status_log("First column of 'Dispatch Instructions' is not datetime. Attempting conversion...")
                try:
                    self.df_dispatch.iloc[:, 0] = pd.to_datetime(self.df_dispatch.iloc[:, 0], errors='coerce')
                    if self.df_dispatch.iloc[:, 0].isnull().all() and not self.df_dispatch.empty:
                         messagebox.showerror("Data Error", "First column of 'Dispatch Instructions' (expected dates) could not be converted to dates/timestamps. All values are invalid.")
                         self.status_log("Error: All values in first column of 'Dispatch Instructions' failed datetime conversion.")
                         self.df_dispatch = None 
                    elif self.df_dispatch.iloc[:, 0].isnull().any():
                        num_failed = self.df_dispatch.iloc[:,0].isnull().sum()
                        self.status_log(f"Warning: {num_failed} values in the first column of 'Dispatch Instructions' failed datetime conversion and were set to NaT.")
                        messagebox.showwarning("Data Conversion Warning", f"{num_failed} date entries in 'Dispatch Instructions' are invalid and were ignored.")
                except Exception as e_conv:
                    messagebox.showerror("Data Error", f"Could not convert first column of 'Dispatch Instructions' to dates/timestamps.\nError: {e_conv}")
                    self.status_log(f"Error converting first column of 'Dispatch Instructions': {e_conv}")
                    self.df_dispatch = None 
            if self.df_dispatch is not None and not self.df_dispatch.empty and pd.api.types.is_datetime64_any_dtype(self.df_dispatch.iloc[:, 0]):
                valid_dates = self.df_dispatch.iloc[:, 0].dropna()
                if not valid_dates.empty:
                    available_months = sorted(valid_dates.dt.strftime('%b-%y').unique().tolist())
                if not available_months: 
                    self.status_log("No valid months found in 'Dispatch Instructions' (Column A after processing).")
            self.df_availability = pd.read_excel(xls, sheet_name="Availability")
            self.status_log(f"'Availability' sheet read. Rows: {len(self.df_availability)}, Columns: {len(self.df_availability.columns) if self.df_availability is not None else 0}")
            if self.df_availability is None or self.df_availability.empty: 
                 self.status_log("Warning: 'Availability' sheet is empty or failed to load.")
                 self.df_availability = None 
            elif self.df_availability.columns.empty:
                self.status_log("Warning: 'Availability' sheet has no columns.")
                messagebox.showwarning("Data Warning", "'Availability' sheet has no columns.")
                self.df_availability = None
            elif len(self.df_availability.columns) < 4: 
                 self.status_log("Warning: 'Availability' sheet has fewer than 4 columns. Required for Final Availability lookup from Column D.")
                 messagebox.showwarning("Data Warning", "'Availability' sheet has fewer than 4 columns. Final Availability features might not work correctly.")
            return self.df_dispatch, self.df_availability, available_months
        except FileNotFoundError:
            messagebox.showerror("File Error", f"File not found: {file_path}")
            self.status_log(f"Error: File not found at {file_path}")
            return None, None, [] 
        except ValueError as ve: 
            messagebox.showerror("File Error", f"Error reading Excel file. It might be corrupted or not a valid Excel file.\nDetails: {ve}")
            self.status_log(f"Error: Could not read Excel file '{file_path}'. Details: {ve}")
            return None, None, []
        except Exception as e:
            import traceback
            messagebox.showerror("Read Error", f"An unexpected error occurred while reading the Excel file: {e}\n\n{traceback.format_exc()}")
            self.status_log(f"Critical Error reading Excel: {e}\n{traceback.format_exc()}")
            return None, None, []

    def get_initial_availability_load(self, selected_month_dt): 
        self.root.after(0, self.status_log_safe, f"Attempting to fetch Initial Hourly Availability for {selected_month_dt.strftime('%b-%y')}.")
        timestamp_first_minute_of_month = selected_month_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        initial_load = self.get_hourly_final_availability(timestamp_first_minute_of_month)
        if initial_load is not None:
            self.root.after(0, self.status_log_safe, f"Initial Hourly Availability load determined: {initial_load} (from Col D of Availability sheet for hour 00)")
            return initial_load
        else:
            self.root.after(0, self.status_log_safe, f"No availability data found for the first hour (00:00-00:59) of {selected_month_dt.strftime('%b-%y')} in Col D for Initial Load.")
            self.root.after(0, self._show_messagebox_safe, "warning", "Data Warning", f"No availability data for the first hour of {selected_month_dt.strftime('%b-%y')} (Col D). Cannot determine Initial Hourly Availability.")
            return None

    def perform_minute_wise_processing(self, dispatch_df_month, availability_df_month_full, selected_month_dt, start_load):
        self.root.after(0, self.status_log_safe, f"Starting minute-wise processing for {selected_month_dt.strftime('%b-%y')} with initial load: {start_load:.2f} MW.")
        month_start_dt = selected_month_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month_start_dt = (month_start_dt + relativedelta(months=1))
        total_minutes_in_month = int((next_month_start_dt - month_start_dt).total_seconds() / 60)
        self.root.after(0, self.status_log_safe, f"Month: {selected_month_dt.strftime('%b-%y')}, Total minutes: {total_minutes_in_month}")
        timestamps = pd.date_range(start=month_start_dt, periods=total_minutes_in_month, freq='min')
        results_list = []
        
        current_load = float(start_load) # Load at the END of the previous minute / START of current minute ts
        
        parsed_instructions = []
        # ... (Instruction parsing logic as before - confirmed to be okay) ...
        if dispatch_df_month is not None and not dispatch_df_month.empty:
            if len(dispatch_df_month.columns) < 6:
                self.root.after(0, self.status_log_safe, "Dispatch Instructions sheet has insufficient columns (expected at least 6 for A-F). Processing without dispatch instructions.")
            else:
                temp_df = dispatch_df_month.copy()
                try:
                    temp_df.iloc[:, 0] = pd.to_datetime(temp_df.iloc[:, 0], errors='coerce') 
                    temp_df.iloc[:, 1] = pd.to_datetime(temp_df.iloc[:, 1], errors='coerce') 
                    temp_df.iloc[:, 2] = pd.to_numeric(temp_df.iloc[:, 2], errors='coerce') 
                    temp_df.iloc[:, 4] = temp_df.iloc[:, 4].astype(str).str.upper().str.strip() 
                    temp_df.iloc[:, 5] = pd.to_numeric(temp_df.iloc[:, 5], errors='coerce') 
                    temp_df.dropna(subset=[temp_df.columns[0], temp_df.columns[5]], inplace=True) 
                    for _, row in temp_df.sort_values(by=temp_df.columns[0]).iterrows():
                        duration = row.iloc[2]
                        if not pd.notna(duration) or duration <= 0:
                            self.root.after(0, self.status_log_safe, f"Warning: Invalid/Missing Ramp Duration (Col C) for instruction at {row.iloc[0]}. Defaulting to 1 minute.")
                            duration = 1.0 
                        parsed_instructions.append({
                            'instr_time': row.iloc[0],
                            'target_time_stamp': row.iloc[1] if pd.notna(row.iloc[1]) else None,
                            'ramp_duration_minutes': duration,
                            'post_ramp_target_type': row.iloc[4],
                            'target_demand_mw': row.iloc[5]
                        })
                except IndexError:
                     self.root.after(0, self.status_log_safe, "Error accessing expected columns (A-F) in Dispatch Instructions. Check sheet structure.")
                except Exception as e_parse:
                     self.root.after(0, self.status_log_safe, f"Error parsing Dispatch Instructions: {e_parse}")
            if not parsed_instructions and not dispatch_df_month.empty : 
                 self.root.after(0, self.status_log_safe, "Warning: Dispatch Instructions found but could not be parsed. Processing as if no instructions.")

        instr_idx = 0
        active_ramp_rate = 0.0
        ramp_end_time = None 
        is_ramping = False
        current_ramp_target_mw = start_load 
        current_post_ramp_type = None 
        
        active_hourly_availability = np.nan 
        is_following_fcbl_directive = self.start_load_type_var.get() == "FinalAvailabilityHourly"
        # Store the target that led to FCBL mode, for fallback
        fallback_fcbl_target_load = start_load if not is_following_fcbl_directive else np.nan


        log_update_interval = max(1, total_minutes_in_month // 100)

        for i, ts in enumerate(timestamps):
            load_for_this_minute = current_load # Load at the start of minute 'ts'

            if ts.minute == 0: # Update hourly availability cache
                hourly_avail_val = self.get_hourly_final_availability(ts)
                active_hourly_availability = float(hourly_avail_val) if pd.notna(hourly_avail_val) else np.nan

            # 1. Apply active FCBL directive first (overrides ramps)
            if is_following_fcbl_directive:
                hourly_avail = self.get_hourly_final_availability(ts) # Re-check for current hour
                if hourly_avail is not None:
                    load_for_this_minute = hourly_avail
                    # active_dispatch_demand remains np.nan (already set or will be by new FCBL instr)
                else: # Fallback if FCBL lookup fails for this hour
                    load_for_this_minute = fallback_fcbl_target_load if pd.notna(fallback_fcbl_target_load) else load_for_this_minute # Use fallback or last known load
                    # active_dispatch_demand remains fallback_fcbl_target_load
                    self.root.after(0, self.status_log_safe, f"    Minute {ts.strftime('%H:%M')}: FCBL mode, but no availability for this hour. Using fallback load: {load_for_this_minute:.2f}")
            
            # 2. Process new instruction (can override FCBL mode or ongoing ramp)
            if instr_idx < len(parsed_instructions) and ts == parsed_instructions[instr_idx]['instr_time']:
                instr = parsed_instructions[instr_idx]
                self.root.after(0, self.status_log_safe, f"  Minute {ts.strftime('%H:%M')}: Instruction at {instr['instr_time']}. Target: {instr['target_demand_mw']:.2f}, Duration: {instr['ramp_duration_minutes']}, TargetTime: {instr['target_time_stamp']}, PostRamp: {instr['post_ramp_target_type']}")
                
                is_following_fcbl_directive = False # New instruction overrides previous FCBL state initially
                is_ramping = False # Assume new instruction might stop previous ramp
                active_ramp_rate = 0.0

                current_ramp_target_mw = instr['target_demand_mw']
                current_post_ramp_type = instr['post_ramp_target_type']
                fallback_fcbl_target_load = current_ramp_target_mw # Store this as potential fallback

                duration = instr['ramp_duration_minutes']
                
                is_instantaneous = pd.notna(instr['target_time_stamp']) and instr['target_time_stamp'] == ts

                if is_instantaneous:
                    load_for_this_minute = current_ramp_target_mw
                    ramp_end_time = ts 
                    self.root.after(0, self.status_log_safe, f"    Instantaneous change to target: {load_for_this_minute:.2f} MW.")
                    # Post-ramp logic applies immediately for instantaneous
                    if current_post_ramp_type == "FCBL":
                        hourly_avail = self.get_hourly_final_availability(ts)
                        if hourly_avail is not None:
                            load_for_this_minute = hourly_avail
                            is_following_fcbl_directive = True
                            self.root.after(0, self.status_log_safe, f"    Instantaneous FCBL. Load set to: {load_for_this_minute:.2f} MW.")
                        else: # FCBL lookup failed, load_for_this_minute is already target
                            self.root.after(0, self.status_log_safe, f"    Instantaneous FCBL, but no availability. Load remains target: {load_for_this_minute:.2f} MW.")
                elif load_for_this_minute == current_ramp_target_mw: # Already at target
                    self.root.after(0, self.status_log_safe, f"    Load {load_for_this_minute:.2f} already at new target. No ramp. Checking post-ramp.")
                    ramp_end_time = ts # Ends now
                    if current_post_ramp_type == "FCBL": # Check post-ramp FCBL
                        hourly_avail = self.get_hourly_final_availability(ts)
                        if hourly_avail is not None: 
                            load_for_this_minute = hourly_avail; is_following_fcbl_directive = True
                            self.root.after(0, self.status_log_safe, f"    FCBL (no ramp). Load set to: {load_for_this_minute:.2f} MW.")
                        else: self.root.after(0, self.status_log_safe, f"    FCBL (no ramp), no availability. Load remains target: {load_for_this_minute:.2f} MW.")
                else: # Start a ramp
                    is_ramping = True
                    if pd.notna(instr['target_time_stamp']) and instr['target_time_stamp'] > ts:
                        ramp_end_time = instr['target_time_stamp']
                        effective_duration_minutes = max(1.0, (ramp_end_time - ts).total_seconds() / 60)
                        if abs(effective_duration_minutes - duration) > 1 and duration > 0:
                             self.root.after(0, self.status_log_safe, f"    Note: Ramp Duration (Col C: {duration} min) differs from time to Target Time Stamp (Col B). Using duration to Target Time Stamp ({effective_duration_minutes:.1f} min).")
                        duration = effective_duration_minutes
                    else:
                        ramp_end_time = ts + datetime.timedelta(minutes=duration)
                        if pd.notna(instr['target_time_stamp']): # Log if target_time_stamp was not usable
                             self.root.after(0, self.status_log_safe, f"    Warning: Target Time Stamp {instr['target_time_stamp']} is not in future or invalid. Using Ramp Duration {duration} min. Calculated Ramp End: {ramp_end_time.strftime('%Y-%m-%d %H:%M')}")
                    active_ramp_rate = (current_ramp_target_mw - load_for_this_minute) / duration
                    self.root.after(0, self.status_log_safe, f"    New ramp started. Rate: {active_ramp_rate:.2f} MW/min towards {current_ramp_target_mw:.2f} by {ramp_end_time.strftime('%Y-%m-%d %H:%M')}.")
                instr_idx += 1
            
            # 3. If not following FCBL and a ramp is active (and wasn't just completed by a new instruction)
            elif is_ramping and ramp_end_time is not None: # No new instruction, but was ramping
                if ts < ramp_end_time: # Ramp continues
                    # load_for_this_minute was already set by previous iteration's end + ramp_rate
                    # current_load for next iter will be load_for_this_minute + active_ramp_rate
                    pass # The load_for_this_minute is already correct due to previous iteration's calculation for current_load
                elif ts == ramp_end_time: # Ramp ends now
                    load_for_this_minute = current_ramp_target_mw # Snap to target for this minute
                    is_ramping = False
                    self.root.after(0, self.status_log_safe, f"  Minute {ts.strftime('%H:%M')}: Ongoing ramp ended. Load at target: {load_for_this_minute:.2f} MW.")
                    if current_post_ramp_type == "FCBL":
                        hourly_avail = self.get_hourly_final_availability(ts)
                        if hourly_avail is not None:
                            load_for_this_minute = hourly_avail
                            is_following_fcbl_directive = True
                            fallback_fcbl_target_load = current_ramp_target_mw # Store this
                            self.root.after(0, self.status_log_safe, f"    Post-ramp FCBL. Load set to: {load_for_this_minute:.2f} MW.")
                        else: # FCBL lookup failed
                            is_following_fcbl_directive = False # Not following if lookup fails
                            self.root.after(0, self.status_log_safe, f"    Post-ramp FCBL, no availability. Load remains target: {load_for_this_minute:.2f} MW.")
                    else: # Not FCBL
                        is_following_fcbl_directive = False
            
            # Record results for minute ts
            
            results_list.append({
                'Date Time Stamp': ts, 
                'Availability': round(active_hourly_availability, 2) if pd.notna(active_hourly_availability) else pd.NA,
                'Load': round(load_for_this_minute, 3), 
                'LPM (30 Min Sum)': pd.NA  # Will be calculated later
            })

            # Prepare current_load for the START of the next minute
            if is_ramping and ramp_end_time is not None and ts < ramp_end_time :
                current_load = load_for_this_minute + active_ramp_rate
            else: # Not ramping, or ramp just ended
                current_load = load_for_this_minute 
        
        # Final progress update and logging
        self.root.after(0, self.update_progress_safe, 100)
        self.root.after(0, self.status_log_safe, "Core minute-wise load calculation loop finished.")
        if not results_list: # Should not happen if timestamps list is generated
            self.root.after(0, self.status_log_safe, "Warning: No results generated from processing loop.")
            return pd.DataFrame(columns=['Date Time Stamp', 'Availability', 'Load', 'LPM (30 Min Sum)'])

        results_df = pd.DataFrame(results_list)
        
        # Add Target Load column E from dispatch instructions and highlight periods
        self.root.after(0, self.status_log_safe, "Adding Target Load column from dispatch instructions...")
        results_df['Target Load'] = ''
        results_df['Highlight_Row'] = False  # Add column to track which rows to highlight
        
        # Process dispatch instructions to create Target Load entries and mark highlighting periods
        for i, instr in enumerate(parsed_instructions):
            instr_time = instr['instr_time']
            target_time = instr['target_time_stamp']
            post_ramp_type = instr['post_ramp_target_type']
            ramp_duration = instr['ramp_duration_minutes']
            
            # Create the target load string
            if post_ramp_type == 'FCBL':
                target_load_text = 'FCBL'
            else:
                target_load_text = f"{instr['target_demand_mw']:.0f} MW"
            
            # Add timestamp if available
            if pd.notna(target_time):
                timestamp_str = target_time.strftime('%d.%b.%y %H:%M')
                target_load_display = f"{target_load_text} @ {timestamp_str}"
                ramp_end_time = target_time
            else:
                target_load_display = target_load_text
                ramp_end_time = instr_time + datetime.timedelta(minutes=ramp_duration)
            
            # Set the target load for the instruction time
            mask = results_df['Date Time Stamp'] == instr_time
            results_df.loc[mask, 'Target Load'] = target_load_display
            
            # Mark all rows from instruction time until target is met for highlighting
            highlight_mask = (results_df['Date Time Stamp'] >= instr_time) & (results_df['Date Time Stamp'] <= ramp_end_time)
            results_df.loc[highlight_mask, 'Highlight_Row'] = True
        
        self.root.after(0, self.status_log_safe, "Calculating specific LPM (30 Min Sum)...")
        if not results_df.empty:
            # Calculate Load Per Minute for LPM calculation only (not stored in final output)
            load_per_minute = results_df['Load'] / 30.0
            lpm_series = pd.Series(load_per_minute.values, index=results_df['Date Time Stamp'])
            
            for idx, row_ts in results_df['Date Time Stamp'].items(): 
                if row_ts.minute == 30:
                    # Sum from minute 1 to 30 of current hour
                    window_start = row_ts.replace(minute=1)
                    window_end = row_ts  # minute 30
                    if window_start >= lpm_series.index.min():
                        sum_val = lpm_series.loc[window_start:window_end].sum()
                        results_df.loc[idx, 'LPM (30 Min Sum)'] = round(sum_val, 5)
                elif row_ts.minute == 0:
                    # Sum from minute 31 to 0 of next hour (i.e., 31-59 of previous hour + 0 of current hour)
                    window_start = row_ts - pd.Timedelta(minutes=29)  # 29 minutes back from 0 = minute 31 of previous hour
                    window_end = row_ts  # minute 0 of current hour
                    if window_start >= lpm_series.index.min():
                        sum_val = lpm_series.loc[window_start:window_end].sum()
                        results_df.loc[idx, 'LPM (30 Min Sum)'] = round(sum_val, 5)
        self.root.after(0, self.status_log_safe, "LPM (30 Min Sum) calculation complete.")
        return results_df

    def create_summary_sheet(self, writer, result_df):
        """Create a summary sheet with half-hourly LPM data"""
        try:
            # Get the first date to determine sheet name
            first_date = result_df['Date Time Stamp'].iloc[0]
            sheet_name = first_date.strftime('%b-%y')
            
            # Filter for half-hourly data (minutes 0 and 30)
            half_hourly_mask = result_df['Date Time Stamp'].dt.minute.isin([0, 30])
            summary_data = result_df[half_hourly_mask].copy()
            
            # Select only Date Time Stamp and LPM (30 Min Sum) columns
            summary_df = summary_data[['Date Time Stamp', 'LPM (30 Min Sum)']].copy()
            
            # Rename columns for summary sheet
            summary_df.columns = ['Half Hourly Date', 'LPM (30 Min Sum)']
            
            # Write to summary sheet
            summary_df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Get the worksheet for formatting
            summary_worksheet = writer.sheets[sheet_name]
            
            # Apply formatting to summary sheet
            header_font = Font(bold=True, color="FFFFFFFF", name='Calibri')
            header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format headers
            for col_num in range(1, len(summary_df.columns) + 1):
                cell = summary_worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Set column widths
            summary_worksheet.column_dimensions['A'].width = 22  # Half Hourly Date
            summary_worksheet.column_dimensions['B'].width = 15  # LPM (30 Min Sum)
            
            # Format Date column as dd.mmm.yy hh:mm
            for row_idx in range(2, summary_worksheet.max_row + 1):
                summary_worksheet[f'A{row_idx}'].number_format = 'dd.mmm.yy hh:mm'
                summary_worksheet[f'B{row_idx}'].number_format = '#,##0.000'
            
            # Freeze panes
            summary_worksheet.freeze_panes = 'A2'
            
            self.root.after(0, self.status_log_safe, f"Summary sheet '{sheet_name}' created with {len(summary_df)} half-hourly records")
            
        except Exception as e:
            self.root.after(0, self.status_log_safe, f"Error creating summary sheet: {e}")

    def save_output_excel(self, result_df, export_path):
        self.root.after(0, self.status_log_safe, f"Attempting to save output to: {export_path}")
        try:
            with pd.ExcelWriter(export_path, engine='openpyxl', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
                cols = ['Date Time Stamp', 'Availability', 'Load', 
                        'Target Load', 'LPM (30 Min Sum)']
                # Only include columns that should be in the final output (exclude Highlight_Row)
                result_df_ordered = result_df[[col for col in cols if col in result_df.columns]]
                
                result_df_ordered.to_excel(writer, index=False, sheet_name="FADL_Calculation")
                
                # Create summary sheet with half-hourly data
                self.create_summary_sheet(writer, result_df)
                
                workbook = writer.book
                worksheet = writer.sheets["FADL_Calculation"]
                worksheet.freeze_panes = 'A2'
                header_font = Font(bold=True, color="FFFFFFFF", name='Calibri') 
                header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid") 
                center_alignment = Alignment(horizontal="center", vertical="center")

                for col_num, column_title in enumerate(result_df_ordered.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    column_letter = get_column_letter(col_num)
                    header_len = len(str(column_title))
                    if not result_df_ordered[column_title].empty:
                        if result_df_ordered[column_title].isna().all():
                            max_data_len = 0
                        else:
                            max_data_len = result_df_ordered[column_title].astype(str).map(len).max()
                    else: max_data_len = 0
                    max_len = max(header_len, int(max_data_len)) 
                    if column_title == 'Date Time Stamp': 
                        adjusted_width = 22 
                    elif column_title == 'Target Load':
                        adjusted_width = 25  # Wider for target load with timestamp
                    else: 
                        adjusted_width = (max_len + 2) * 1.1 
                    worksheet.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 50)

                # Apply formatting to all columns
                for col_idx, col_name in enumerate(result_df_ordered.columns):
                    col_letter = get_column_letter(col_idx + 1)
                    
                    if col_name == 'Date Time Stamp':
                        # Format Date Time Stamp column as dd.mmm.yy hh:mm
                        for row_idx in range(2, worksheet.max_row + 1):
                            worksheet[f'{col_letter}{row_idx}'].number_format = 'dd.mmm.yy hh:mm'
                    elif col_name != 'Target Load' and pd.api.types.is_numeric_dtype(result_df_ordered[col_name]):
                        # Apply number formatting to numeric columns
                        num_format = '#,##0.000' 
                        if "Sum" in col_name : num_format = '#,##0.000'
                        elif "Load" == col_name or "Availability" == col_name:
                            num_format = '#,##0.00'
                        for row_idx in range(2, worksheet.max_row + 1):
                            worksheet[f'{col_letter}{row_idx}'].number_format = num_format
                
                # Highlight rows during dispatch instruction periods (from instruction until target met)
                target_load_fill = PatternFill(start_color="FFFFCD", end_color="FFFFCD", fill_type="solid")  # RGB 255,255,205
                
                # Check if we have the Highlight_Row column in the original dataframe
                if 'Highlight_Row' in result_df.columns:
                    highlight_data = result_df['Highlight_Row'].values
                    for row_idx in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
                        df_row_idx = row_idx - 2  # Convert Excel row to DataFrame index (Excel row 2 = df index 0)
                        if df_row_idx < len(highlight_data) and highlight_data[df_row_idx]:
                            # Highlight columns A through E (Date Time Stamp, Availability, Load, Target Load, LPM)
                            for col_to_highlight in range(1, 6):  # Columns A=1, B=2, C=3, D=4, E=5
                                cell_to_highlight = worksheet.cell(row=row_idx, column=col_to_highlight)
                                cell_to_highlight.fill = target_load_fill
                
            self.root.after(0, self.status_log_safe, f"Output successfully saved and formatted: {export_path}")
            self.root.after(0, lambda: messagebox.showinfo("Success", f"Processing complete. Output saved to:\n{export_path}"))
            # Open the file automatically
            self.root.after(0, lambda: self.open_file(export_path))
        except PermissionError:
            err_msg = f"Permission denied to save to {export_path}. Check permissions/if file is open."
            self.root.after(0, self.status_log_safe, err_msg)
            self.root.after(0, lambda: messagebox.showerror("Save Error", err_msg))
        except Exception as e:
            import traceback
            err_msg = f"Failed to save output file.\nError: {e}" 
            self.root.after(0, self.status_log_safe, f"Error saving output Excel: {e}\n{traceback.format_exc()}") 
            self.root.after(0, lambda: messagebox.showerror("Save Error", err_msg))

    def update_progress_safe(self, value):
        self.progress_bar['value'] = value

    def status_log_safe(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _show_messagebox_safe(self, msg_type, title, message):
        if msg_type == "info": messagebox.showinfo(title, message)
        elif msg_type == "warning": messagebox.showwarning(title, message)
        elif msg_type == "error": messagebox.showerror(title, message)

    def _processing_logic(self):
        import traceback 
        try:
            self.root.after(0, self.status_log_safe, "Background processing thread started.")
            selected_month_str = self.month_var.get()
            custom_load_val = self.custom_load_var.get() 
            export_dir = self.export_dir_var.get()
            start_load_type = self.start_load_type_var.get()

            try:
                selected_month_dt = datetime.datetime.strptime(selected_month_str, '%b-%y')
            except ValueError:
                self.root.after(0, self._show_messagebox_safe, "error", "Date Error", f"Invalid month format: {selected_month_str}.")
                self.root.after(0, self.status_log_safe, f"Processing aborted: Invalid month format '{selected_month_str}'.")
                return 

            start_load = 0.0
            if start_load_type == "FinalAvailabilityHourly": 
                start_load = self.get_initial_availability_load(selected_month_dt) 
                if start_load is None:
                    self.root.after(0, self.status_log_safe, "Processing aborted: Could not determine Initial Hourly Availability for starting load.")
                    return 
            else: # Custom
                start_load = custom_load_val 
            
            self.root.after(0, self.status_log_safe, f"Determined start load: {start_load:.2f} MW")

            month_start_filter = selected_month_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            month_end_filter = (month_start_filter + relativedelta(months=1))
            
            dispatch_df_month = pd.DataFrame() 
            if self.df_dispatch is not None and not self.df_dispatch.empty:
                if not pd.api.types.is_datetime64_any_dtype(self.df_dispatch.iloc[:, 0]):
                    self.root.after(0, self.status_log_safe, "Warning: Re-converting Dispatch timestamp column in thread before filtering.")
                    try: self.df_dispatch.iloc[:, 0] = pd.to_datetime(self.df_dispatch.iloc[:, 0], errors='coerce')
                    except: self.root.after(0, self.status_log_safe, "Error during defensive conversion of Dispatch timestamp.")
                
                if pd.api.types.is_datetime64_any_dtype(self.df_dispatch.iloc[:, 0]): 
                    dispatch_df_month = self.df_dispatch[(self.df_dispatch.iloc[:, 0] >= month_start_filter) & (self.df_dispatch.iloc[:, 0] < month_end_filter)].copy()
            
            if self.df_availability is not None and not self.df_availability.empty and self.availability_series_hourly_lookup is None:
                 self.root.after(0, self.status_log_safe, "Warning: Hourly availability lookup was not prepared. Attempting now.")
                 self._prepare_hourly_availability_lookup() 

            self.root.after(0, self.status_log_safe, f"Filtered Dispatch Data for {selected_month_str}: {len(dispatch_df_month)} rows.")
            
            if dispatch_df_month.empty and start_load_type != "Custom" and (start_load == 0 or start_load is None) : 
                 self.root.after(0, self.status_log_safe, f"Warning: No dispatch instructions for {selected_month_str} and starting load is 0 or could not be determined from availability.")
        
            result_df = self.perform_minute_wise_processing(dispatch_df_month, self.df_availability, selected_month_dt, start_load) 
            
            output_file_name = "FADL Calculation.xlsx"
            import os 
            full_export_path = os.path.join(export_dir, output_file_name)
            self.save_output_excel(result_df, full_export_path) 
            self.root.after(0, self.status_log_safe, "Processing thread finished successfully.")

        except Exception as e:
            error_message = f"Critical error in processing thread: {e}\n{traceback.format_exc()}"
            self.root.after(0, self.status_log_safe, error_message)
            self.root.after(0, self._show_messagebox_safe, "error", "Processing Error", f"An critical unexpected error occurred: {e}")
        finally:
            self.root.after(0, lambda: self.process_button.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.progress_bar.config(value=0)) 
            self.root.after(0, self.status_log_safe, "Background processing thread has ended.")

    def start_processing_thread(self):
        self.status_log("Start processing button clicked. Validating inputs...")
        if self.df_dispatch is None : 
            messagebox.showerror("Input Error", "Dispatch data not loaded. Please select a valid Excel file with a 'Dispatch Instructions' sheet.")
            self.status_log("Processing aborted: Dispatch data not loaded.")
            return
        if self.start_load_type_var.get() == "FinalAvailabilityHourly" and self.df_availability is None:
            messagebox.showerror("Input Error", "Availability data not loaded. This is required for 'Use Final Availability' starting load option.")
            self.status_log("Processing aborted: Availability data required for selected start type but not loaded.")
            return

        selected_month_str = self.month_var.get()
        if not selected_month_str or selected_month_str in ["Select a file first", "Error reading file", "No months found", "Error in sheet data", "Error reading/parsing file"]:
            messagebox.showerror("Input Error", "Please select a valid month to process from a successfully loaded file.")
            self.status_log("Processing aborted: Month not selected or invalid.")
            return

        if self.start_load_type_var.get() == "Custom":
            try:
                custom_load_val = self.custom_load_var.get() 
                if custom_load_val < 0:
                    messagebox.showerror("Input Error", "Custom load cannot be negative.")
                    self.status_log("Processing aborted: Negative custom load.")
                    return
            except tk.TclError: 
                messagebox.showerror("Input Error", "Invalid custom load value. Please enter a number.")
                self.status_log("Processing aborted: Invalid custom load value format.")
                return
        
        export_dir_path = self.export_dir_var.get()
        if not export_dir_path:
            messagebox.showerror("Input Error", "Please select an export directory.")
            self.status_log("Processing aborted: Export directory not selected.")
            return
        
        import os 
        if not os.path.isdir(export_dir_path):
            messagebox.showerror("Input Error", f"The selected export path is not a valid directory:\n{export_dir_path}")
            self.status_log(f"Processing aborted: Export path '{export_dir_path}' is not a directory.")
            return

        self.status_log("Inputs validated. Starting background processing task...")
        self.process_button.config(state=tk.DISABLED)
        self.progress_bar['value'] = 0
        self.root.update_idletasks() 

        processing_thread = threading.Thread(target=self._processing_logic, daemon=True)
        processing_thread.start()

    def open_file(self, file_path):
        """Open the Excel file with the default application"""
        try:
            if sys.platform.startswith('win'):  # Windows
                os.startfile(file_path)
                self.status_log_safe(f"Opened file: {os.path.basename(file_path)}")
            elif sys.platform.startswith('darwin'):  # macOS
                subprocess.call(['open', file_path])
                self.status_log_safe(f"Opened file: {os.path.basename(file_path)}")
            else:  # Linux and other Unix-like systems
                subprocess.call(['xdg-open', file_path])
                self.status_log_safe(f"Opened file: {os.path.basename(file_path)}")
        except Exception as e:
            self.status_log_safe(f"Could not open file automatically: {e}")
            self.status_log_safe(f"File location: {file_path}")

    def status_log(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{datetime.datetime.now().strftime('%d.%b.%y %H:%M')} - {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        print(message)

if __name__ == "__main__":
    root = tk.Tk()
    app = LoadProcessorApp(root)
    root.mainloop()
